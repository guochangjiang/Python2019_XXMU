# create_xlsx.py

import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Alignment, Font

# 设置单元格格式，其中字体及背景为默认
style0 = NamedStyle('style0')
style0.border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
style0.alignment = Alignment(horizontal='center',vertical='center',)
font0 = Font(bold=True)

# 生成工作簿及sheet
f = openpyxl.Workbook()
table = f.active
table.title = '成绩单'
table.sheet_properties.tabColor = 'FF0000'

# 信息列表
subject_list = ['语文', '英语', '政治', '历史', '地理','数学', '物理', '化学', '生物']
info_list = [
['同学A',201901,80,85,90,77,91,88,85,92,100],
['同学B',201902,91,88,95,90,95,99,89,100,90],
['同学C',201903,75,70,98,100,77,92,84,99,95],
['同学D',201904,85,90,93,87,97,96,94,89,98]
]

# 创建表头并写入
table.merge_cells('A1:A2')
table.cell(row = 1,column = 1,value = '姓名')
table.merge_cells('B1:B2')
table.cell(row = 1,column = 2,value = '学号')
table.merge_cells('C1:G1')
table.cell(row = 1,column = 3,value = '文科')
table.merge_cells('H1:K1')
table.cell(row = 1,column = 8,value = '理科')
for i in range(len(subject_list)):
    table.cell(row = 2,column = i+3,value = subject_list[i])

# 写入信息
for obs in range(len(info_list)):
    for info in range(len(info_list[obs])):
        table.cell(row = obs+3,column = info+1,value = info_list[obs][info])

# 格式应用
for row in table.rows:
    for cell in row:
        cell.style = style0
for cell in list(table.rows)[0]:
    cell.font=font0

# 保存文件
f.save("score.xlsx")