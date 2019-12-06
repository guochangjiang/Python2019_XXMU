# read_xlsx
import openpyxl

#load_workbook()函数来加载 Excel 文档
wb = openpyxl.load_workbook('ProgramLang.xlsx')
#读取 Excel 工作表（sheet）
print('取得所有工作表的表名 :')
print(wb.sheetnames)
print('取得某张工作表 :')
sheet = wb['PYPL']
print('表名 - ' + sheet.title)
print('取得活动工作表 :')
active_sheet = wb.active
print('表名 - ' + active_sheet.title)

# 获取工作表大小
print('获取工作表的大小:')
print('总行数 -> ' + str(active_sheet.max_row))
print('总列数 -> ' + str(active_sheet.max_column))

# 读取单元格 （Cell）
print('取得 A1 单元格 :')
cell = active_sheet['A1']
print(cell.value)

print('取得 B1 单元格 :')
cell = active_sheet['B1']
print(cell.value)

print('行号为 ' + str(cell.row) + ',列号为 ' + str(cell.column) + ' 的单元格，其值为 ' + cell.value)
print('单元格 ' + cell.coordinate + ' 其值为 ' + cell.value)

print('取得 C1 单元格的值 :')
print(active_sheet['C1'].value)

print('通过指定行与列，来获取单元格:')
print(active_sheet.cell(row=1, column=2).value)

print('迭代行与列，来获取单元格的值:')
for i in range(1, 11):
    print(i, active_sheet.cell(row=i, column=2).value)

# 列转换函数
## openpyxl 提供了两个函数，用于转换列号：
## openpyxl.utils.get_column_letter -> 会把数字转化为字母。
## openpyxl.utils.column_index_from_string -> 会把字母转化为数字。
from openpyxl.utils import get_column_letter, column_index_from_string
print('列转换函数:')
print('[数字转换为字母]')
print('第 1 列 -> ' + get_column_letter(1))
print('第 2 列 -> ' + get_column_letter(2))
print('第 37 列 -> ' + get_column_letter(37))
print('第 818 列 -> ' + get_column_letter(818))
print('[字母转换为数字]')
print('第 A 列 -> ' + str(column_index_from_string('A')))
print('第 CC 列 -> ' + str(column_index_from_string('CC')))

# 切片
## 对 Worksheet 对象切片，取得表格中的一个矩形区域，迭代遍历这个区域中的所有 Cell 对象。

print(tuple(active_sheet['A2':'D4']))
for row_objects in active_sheet['A2':'D4']:
    for cell_object in row_objects:
        print(cell_object.coordinate, cell_object.value)
    print('-- 当前行获取结束 --')

# 获取指定行或指定列
## Worksheet 对象的 rows 和 columns 属性，来获取指定行或者列
print('获取特定行:')
print(list(active_sheet.rows)[2])
for cell_object in list(active_sheet.rows)[2]:
    print(cell_object.value)

print('获取特定列:')
print(list(active_sheet.columns)[2])
for cell_object in list(active_sheet.columns)[2]:
    print(cell_object.value)